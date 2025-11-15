import configparser
from openpyxl import load_workbook
import os

class ExcelColumnReader:
    errors = {
        -1000: "Calibration file does not exist",
        -1001: "Sheet does not exist",
        -1002: "No parameters found",
        -1005: "Wrong config file",
        - 1006: "Wrong config file keys",
    }


    def __init__(self, config_file="config.ini"):
        """
        Initialize with config file path.
        Reads Excel file and sheet name from config.

        Possible errors
        -1000: file does not exist
        -1001: sheet does not exist
        -1002: no parameters found
        -1005: wrong config file
        """
        self.config = configparser.ConfigParser()
        if not os.path.exists(config_file):
            self.file_path = None
            self.sheet_name = None
            self.error = -1005
            return

        self.config.read(config_file)

        try:
            self.file_path = self.config.get("calibration_data", "file")
            self.sheet_name = self.config.get("calibration_data", "spreadsheet")
        except (configparser.NoSectionError, configparser.NoOptionError):
            self.file_path = None
            self.sheet_name = None
            self.error = -1006
            return

        if not os.path.exists(self.file_path):
            self.error = -1000
            return

        self.error = 0  # no error
        self.workbook = load_workbook(self.file_path, data_only=True)

    def read_column(self, column="A"):
        """
        Read all values from the specified column of the configured sheet.
        Returns a list of non-empty cell values, or error codes:

        -1000: file does not exist
        -1001: sheet does not exist
        -1002: no parameters found
        -1005: wrong config file
        """
        if hasattr(self, "error") and self.error != 0:
            return self.error

        if self.sheet_name not in self.workbook.sheetnames:
            return -1001

        sheet = self.workbook[self.sheet_name]
        column_values = [row.value for row in sheet[column] if row.value is not None]

        if not column_values:
            return -1002

        return column_values

    def get_error_text(self, error_num):
        return ExcelColumnReader.errors[error_num]

